from __future__ import annotations

import argparse
import datetime as _dt
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook


COLUMN_MAP = {
    "人名（用#作为注释）": "comment",
    "first_name": "first_name",
    "last_name": "last_name",
    "dynasty": "dynasty",
    "culture": "culture",
    "religion": "religion",
    "birth_date": "birth_date",
    "birth": "birth",
    "death_date": "death_date",
    "father": "father",
    "tag": "tag",
    "adm": "adm",
    "dip": "dip",
    "mil": "mil",
    "script": "script",
    "备注": "note",
}

FIELD_ORDER = [
    "first_name",
    "last_name",
    "culture",
    "religion",
    "adm",
    "dip",
    "mil",
    "birth_date",
    "death_date",
    "birth",
    "dynasty",
    "father",
    "tag",
]


def cast_scalar(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (int, bool)):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, (_dt.datetime, _dt.date)):
        return f"{value.year}.{value.month}.{value.day}"
    text = str(value).strip()
    return text or None


def render_field(key: str, value: str) -> Optional[str]:
    if value is None:
        return None
    if key in {"first_name", "last_name"}:
        return f"\t\t{key} = {{ name = {value} }}"
    if key in {"adm", "dip", "mil"}:
        return f"\t\t{key} = {value}"
    return f"\t\t{key} = {value}"


def render_entry(record: Dict[str, Optional[str]]) -> str:
    lines: List[str] = []
    comment = record.get("comment")
    note = record.get("note")
    if comment:
        lines.append(f"\t# {comment}")
    if note:
        lines.append(f"\t# 备注：{note}")
    identifier = record.get("identifier")
    if not identifier:
        raise ValueError("记录缺少 identifier，无法生成条目。")
    lines.append(f"\t{identifier} = {{")
    for field in FIELD_ORDER:
        rendered = render_field(field, record.get(field))
        if rendered:
            lines.append(rendered)
    lines.append("\t}\n")
    return "\n".join(lines)


def prepare_records(rows: Iterable[Iterable[Any]]) -> List[Dict[str, Optional[str]]]:
    rows = iter(rows)
    headers = [COLUMN_MAP.get(h, None) for h in next(rows)]
    records: List[Dict[str, Optional[str]]] = []
    for raw_row in rows:
        record: Dict[str, Optional[str]] = {}
        for header, value in zip(headers, raw_row):
            if header is None:
                continue
            record[header] = cast_scalar(value)
        if not record.get("tag"):
            continue
        records.append(record)
    return records


def normalize_slug(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    slug = text.strip().lower()
    if slug.startswith("name_"):
        slug = slug[len("name_") :]
    slug = slug.replace(" ", "_")
    cleaned = []
    for ch in slug:
        if ch.isalnum() or ch == "_":
            cleaned.append(ch)
        else:
            cleaned.append("_")
    slug = "".join(cleaned).strip("_")
    return slug or None


def build_identifier(
    record: Dict[str, Optional[str]], index: int, slug_counter: Dict[str, int]
) -> str:
    tag = record.get("tag")
    if not tag:
        raise ValueError(f"第 {index} 行缺少 tag，无法生成标识符。")
    base_slug = (
        normalize_slug(record.get("first_name"))
        or normalize_slug(record.get("comment"))
        or f"char_{index:03d}"
    )
    key_base = f"{tag.lower()}_1644_{base_slug}"
    count = slug_counter.get(key_base, 0)
    slug_counter[key_base] = count + 1
    if count == 0:
        return key_base
    return f"{key_base}_{count}"


def main() -> None:
    parser = argparse.ArgumentParser(description="将 1644 角色表转换为 character_db 定义。")
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("1644_characters.xlsx"),
        help="Excel 数据源路径（默认：1644_characters.xlsx）",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="要读取的工作表名称，缺省则使用第一个工作表。",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="输出文件路径；缺省则写到标准输出。",
    )
    args = parser.parse_args()

    if not args.input.exists():
        parser.error(f"找不到输入文件：{args.input}")

    workbook = load_workbook(args.input, data_only=True)
    worksheet = workbook[args.sheet] if args.sheet else workbook.active

    records = prepare_records(worksheet.iter_rows(values_only=True))
    slug_counter: Dict[str, int] = {}
    for idx, record in enumerate(records, start=1):
        record["identifier"] = build_identifier(record, idx, slug_counter)
    rendered_entries = "\n".join(render_entry(record) for record in records)
    output_text = "character_db={\n" + rendered_entries + "}\n"

    if args.output:
        args.output.write_text(output_text, encoding="utf-8")
    else:
        sys.stdout.write(output_text)


if __name__ == "__main__":
    main()

