from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


EXCEL_CANDIDATES: Tuple[Path, ...] = (
    Path("docs") / "eu5_1644人物三围 v0.2.xlsx",
    Path("docs") / "eu5_1644人物三围 v0.1.xlsx",
    Path("eu5_1644人物三围 v0.2.xlsx"),
    Path("eu5_1644人物三围 v0.1.xlsx"),
    Path("1644_characters.xlsx"),
)

TARGET_CANDIDATES: Tuple[Path, ...] = (
    Path("main_menu") / "setup" / "start" / "99_1644_05_characters.txt",
    Path("main_menu") / "setup" / "start" / "zzz_05_characters.txt",
)

HEADER_MAP = {
    "adm(行政)": "adm",
    "dip（外交）": "dip",
    "mil（军事）": "mil",
    "script": "script",
    "备注": "note",
}


@dataclass(frozen=True)
class StatBlock:
    adm: Optional[int]
    dip: Optional[int]
    mil: Optional[int]
    note: Optional[str]

    @classmethod
    def from_row(cls, row: Sequence[Optional[object]], indexes: Dict[str, int]) -> Optional["StatBlock"]:
        def pull(key: str) -> Optional[int]:
            idx = indexes.get(key)
            if idx is None:
                return None
            raw = row[idx]
            if raw is None:
                return None
            if isinstance(raw, (int, float)):
                return int(raw)
            text = str(raw).strip()
            if not text:
                return None
            return int(float(text))

        adm = pull("adm")
        dip = pull("dip")
        mil = pull("mil")
        note_value: Optional[str] = None
        note_idx = indexes.get("note")
        if note_idx is not None:
            raw_note = row[note_idx]
            if raw_note is not None:
                note_value = str(raw_note).strip() or None

        if adm is None and dip is None and mil is None:
            return None
        return cls(adm=adm, dip=dip, mil=mil, note=note_value)


def resolve_workbook_path(path_hint: Optional[str]) -> Path:
    if path_hint:
        candidate = Path(path_hint)
        if not candidate.exists():
            raise FileNotFoundError(f"指定的 Excel 文件不存在：{candidate}")
        return candidate
    for candidate in EXCEL_CANDIDATES:
        if candidate.exists():
            return candidate
    raise FileNotFoundError("未找到任何可用的 Excel 源文件，请使用 --input 参数指定。")


def resolve_target_path(path_hint: Optional[str]) -> Path:
    if path_hint:
        candidate = Path(path_hint)
        if not candidate.exists():
            raise FileNotFoundError(f"指定的目标脚本不存在：{candidate}")
        return candidate
    for candidate in TARGET_CANDIDATES:
        if candidate.exists():
            return candidate
    raise FileNotFoundError("未找到目标角色脚本，请使用 --target 参数指定。")


def load_stats(path_hint: Optional[str]) -> Tuple[Dict[str, StatBlock], Path]:
    workbook_path = resolve_workbook_path(path_hint)
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 3:
        raise RuntimeError("Excel 数据不足，无法解析。")

    header_row = rows[1]
    indexes: Dict[str, int] = {}
    for idx, header in enumerate(header_row):
        if header is None:
            continue
        canonical = HEADER_MAP.get(str(header).strip())
        if canonical:
            indexes[canonical] = idx

    for field in ("script", "adm", "dip", "mil"):
        if field not in indexes:
            raise RuntimeError(f"Excel 缺少必要列：{field}")

    mapping: Dict[str, StatBlock] = {}
    for row in rows[2:]:
        script_cell = row[indexes["script"]]
        if script_cell is None:
            continue
        identifier = str(script_cell).strip()
        if not identifier:
            continue
        stats = StatBlock.from_row(row, indexes)
        if stats is None:
            continue
        mapping[identifier] = stats

    return mapping, workbook_path


def process_block(block_lines: List[str], stats: StatBlock) -> List[str]:
    cleaned: List[str] = []
    for line in block_lines:
        stripped = line.strip()
        if stripped.startswith("adm =") or stripped.startswith("dip =") or stripped.startswith("mil ="):
            continue
        if stripped.startswith("# 三围备注："):
            continue
        cleaned.append(line)

    if not cleaned:
        return block_lines

    indent_block = cleaned[0][: len(cleaned[0]) - len(cleaned[0].lstrip())]
    field_indent = indent_block + "\t"

    stat_entries: List[Tuple[str, Optional[int]]] = [
        ("adm", stats.adm),
        ("dip", stats.dip),
        ("mil", stats.mil),
    ]
    stat_lines = [
        f"{field_indent}{key} = {value}"
        for key, value in stat_entries
        if value is not None
    ]

    if not stat_lines:
        return cleaned

    insert_pos = len(cleaned) - 1  # 默认在闭合大括号之前
    for idx, line in enumerate(cleaned):
        stripped = line.strip()
        if stripped.startswith("religion ="):
            insert_pos = idx + 1
    insert_pos = min(insert_pos, len(cleaned) - 1)

    cleaned[insert_pos:insert_pos] = stat_lines
    if stats.note:
        comment_line = f"{field_indent}# 三围备注：{stats.note}"
        cleaned.insert(insert_pos + len(stat_lines), comment_line)
    return cleaned


def inject_version_comment(lines: List[str], version_label: str) -> List[str]:
    version_comment = f"# 三围数据版本：{version_label}"
    if any(line.strip() == version_comment for line in lines):
        return lines

    insert_idx = 0
    for idx, line in enumerate(lines):
        if line.strip().startswith("character_db"):
            insert_idx = idx
            break
    return lines[:insert_idx] + [version_comment] + lines[insert_idx:]


def rewrite_character_file(stats_map: Dict[str, StatBlock], target_path: Path, version_label: str) -> None:
    original_lines = target_path.read_text(encoding="utf-8").splitlines()
    result: List[str] = []
    i = 0

    while i < len(original_lines):
        line = original_lines[i]
        stripped = line.strip()
        if not stripped.startswith("chi_") or not stripped.endswith("= {"):
            result.append(line)
            i += 1
            continue

        identifier = stripped.split("=")[0].strip()
        block: List[str] = [line]
        brace_depth = line.count("{") - line.count("}")
        i += 1
        while i < len(original_lines) and brace_depth > 0:
            block_line = original_lines[i]
            block.append(block_line)
            brace_depth += block_line.count("{")
            brace_depth -= block_line.count("}")
            i += 1

        stats = stats_map.get(identifier)
        if stats:
            processed = process_block(block, stats)
            result.extend(processed)
        else:
            result.extend(block)

    result = inject_version_comment(result, version_label)
    target_path.write_text("\n".join(result) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="根据 Excel 三围数据更新 99_1644_05_characters.txt。")
    parser.add_argument("--input", help="Excel 数据源路径；缺省时自动搜索。")
    parser.add_argument("--target", help="要写入的角色脚本路径；缺省时自动搜索。")
    args = parser.parse_args()

    stats_map, workbook_path = load_stats(args.input)
    target_path = resolve_target_path(args.target)
    version_label = workbook_path.name
    rewrite_character_file(stats_map, target_path, version_label)


if __name__ == "__main__":
    main()


